from openpyxl import load_workbook
import openpyxl
from tkinter import *
from tkinter import filedialog
import os

batcha=[]
batchb=[]
batchc=[]
batchd=[]
batche= []
batchf= []

wb = load_workbook('Namelist.xlsx')
ws = wb['Sheet1']
batchacol = ws['A']
batchbcol = ws['B']
batchccol = ws['C']
batchdcol = ws['D']
batchecol = ws['E']
batchfcol = ws['F']

c=0
for i in batchacol:
    if c != 0:
        batcha.append(i.value)
    else:
        c = c + 1

c=0
for i in batchbcol:
    if c != 0:
        batchb.append(i.value)
    else:
        c = c + 1

c=0
for i in batchccol:
    if c != 0:
        batchc.append(i.value)
    else:
        c = c + 1

c=0
for i in batchdcol:
    if c != 0:
        batchd.append(i.value)
    else:
        c = c + 1

c=0
for i in batchecol:
    if c != 0:
        batche.append(i.value)
    else:
        c = c + 1

c=0
for i in batchfcol:
    if c != 0:
        batchf.append(i.value)
    else:
        c = c + 1

root = Tk()
root.title('Selected the attedance file from teams:')
root.iconbitmap('logos.ico')
root.filename = filedialog.askopenfilename()

filename = os.path.basename(root.filename)
sheetname = os.path.splitext(filename)[0]
Label(root, text = "Selected the file: "+filename+". Please close this window to proceed").pack()


wb = load_workbook("working/"+filename)

ws= wb[sheetname]
root.mainloop()

print("Working on the sheet:  "+str(ws.title))

col1 = ws['A']
for data in col1:
    print(data.value)

## Removing the duplicates
newcol= []
for data in col1:
    if data.value not in newcol:
        newcol.append(data.value)

print("After removing reduncies:   \n")
col1 = ws['A']

# First years
countbatcha = 0
countbatchb = 0
countbatchc = 0
countbatchd = 0
countbatche = 0
countbatchf = 0

others = []
for data in newcol:
    print(data)
    if data in batcha:
        countbatcha = countbatcha + 1
    elif data in batchb:
        countbatchb = countbatchb + 1
    elif data in batchc:
        countbatchc = countbatchc + 1
    elif data in batchd:
        countbatchd = countbatchd + 1
    elif data in batche:
        countbatche = countbatche + 1
    elif data in batchf:
        countbatchf = countbatchf + 1
    else:
        others.append(data)

counttot = countbatcha+countbatchb+countbatchc+countbatchd+countbatche+countbatchf;

print("Others who are not in list: ")
for data in others:
    print(data)

date= input("Enter the date: ")

attbook = load_workbook("Attendance.xlsx")
attsheet = attbook['Sheet1']

count = attsheet['A2'].value

attsheet['A'+str(count+5)]= count+1
attsheet['B'+str(count+5)]= date
attsheet['C'+str(count+5)]= countbatcha
attsheet['D'+str(count+5)]= countbatchb
attsheet['E'+str(count+5)]= countbatchc
attsheet['F'+str(count+5)]= countbatchd
attsheet['G'+str(count+5)]= countbatche
attsheet['H'+str(count+5)]= countbatchf
attsheet['I'+str(count+5)]= counttot

attsheet['A2']=count + 1

attbook.save("Attendance.xlsx")

check =  load_workbook("Attendance.xlsx")
attsheet = check['Sheet1']
print("Total no.of session till now: "+ str(attsheet['A2'].value))
print("Thank you !!!")
